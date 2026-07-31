#!/usr/bin/env python3
"""Convert a primer-design Excel sheet to an iVar-compatible BED file."""

from __future__ import annotations

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


IUPAC_BASES = set("ACGTRYSWKMBDHVN")
COMPLEMENT = str.maketrans("ACGTRYSWKMBDHVN", "TGCAYRSWMKVHDBN")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="将引物设计 Excel 转换为 iVar/ARTIC 可用的六列 BED。"
    )
    parser.add_argument("input_xlsx", type=Path, help="输入 .xlsx 文件")
    parser.add_argument("output_bed", type=Path, help="输出六列 BED 文件")
    parser.add_argument("--sheet", help="工作表名称；默认使用第一个工作表")
    parser.add_argument(
        "--audit-tsv",
        type=Path,
        help="可选：输出包含 primer 序列和参考匹配结果的审计表",
    )
    parser.add_argument(
        "--reference",
        type=Path,
        help="可选：MN908947.3 FASTA，用于校验坐标和序列",
    )
    parser.add_argument(
        "--strict-reference",
        action="store_true",
        help="存在参考序列 mismatch 时退出；变异适配引物通常不要开启",
    )
    parser.add_argument(
        "--repair-one-based-inclusive",
        action="store_true",
        help="仅当参考序列精确匹配时，将疑似 1-based inclusive 行的 start 减 1",
    )
    return parser.parse_args()


def normalize_sequence(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def reverse_complement(sequence: str) -> str:
    return sequence.translate(COMPLEMENT)[::-1]


def read_fasta(path: Path) -> dict[str, str]:
    records: dict[str, list[str]] = {}
    current_name: str | None = None
    with path.open(encoding="utf-8") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith(">"):
                current_name = line[1:].split()[0]
                if current_name in records:
                    raise ValueError(f"FASTA 中存在重复序列名称: {current_name}")
                records[current_name] = []
            elif current_name is None:
                raise ValueError("FASTA 第一条序列之前缺少 > 标题行")
            else:
                records[current_name].append(line.upper())
    return {name: "".join(parts) for name, parts in records.items()}


def is_header(row: tuple[object, ...]) -> bool:
    cells = [str(value or "").strip().lower() for value in row[:7]]
    return bool(
        set(cells)
        & {
            "chrom",
            "chromosome",
            "reference",
            "start",
            "end",
            "name",
            "strand",
            "sequence",
            "序列",
            "起始",
            "终止",
        }
    )


def coerce_coordinate(value: object, label: str, excel_row: int) -> int:
    if isinstance(value, bool):
        raise ValueError(f"第 {excel_row} 行 {label} 不是整数: {value!r}")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"第 {excel_row} 行 {label} 不是数字: {value!r}") from exc
    if not number.is_integer():
        raise ValueError(f"第 {excel_row} 行 {label} 不是整数: {value!r}")
    return int(number)


def count_mismatches(primer: str, target: str) -> int:
    # N 和其他 IUPAC 模糊碱基按兼容位点处理，避免把简并引物误判为错误。
    compatibility = {
        "A": set("A"),
        "C": set("C"),
        "G": set("G"),
        "T": set("T"),
        "R": set("AG"),
        "Y": set("CT"),
        "S": set("GC"),
        "W": set("AT"),
        "K": set("GT"),
        "M": set("AC"),
        "B": set("CGT"),
        "D": set("AGT"),
        "H": set("ACT"),
        "V": set("ACG"),
        "N": set("ACGT"),
    }
    return sum(base not in compatibility[code] for code, base in zip(primer, target))


def main() -> int:
    args = parse_args()
    workbook = load_workbook(args.input_xlsx, read_only=True, data_only=True)
    if args.sheet:
        if args.sheet not in workbook.sheetnames:
            raise ValueError(
                f"找不到工作表 {args.sheet!r}；可用工作表: {', '.join(workbook.sheetnames)}"
            )
        sheet = workbook[args.sheet]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    reference = read_fasta(args.reference) if args.reference else None
    if args.repair_one_based_inclusive and reference is None:
        raise ValueError("--repair-one-based-inclusive 必须同时提供 --reference")
    records: list[dict[str, object]] = []
    seen_names: set[str] = set()

    for excel_row, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not any(value not in (None, "") for value in row):
            continue
        if not records and is_header(row):
            continue
        if len(row) < 7:
            raise ValueError(f"第 {excel_row} 行不足 7 列")

        chrom = str(row[0] or "").strip()
        start = coerce_coordinate(row[1], "start", excel_row)
        end = coerce_coordinate(row[2], "end", excel_row)
        input_start, input_end = start, end
        name = str(row[3] or "").strip()
        pool = str(row[4] or "").strip()
        strand = str(row[5] or "").strip()
        sequence = normalize_sequence(row[6])

        if not chrom or not name or not pool or not sequence:
            raise ValueError(f"第 {excel_row} 行存在必填空值")
        if start < 0 or end <= start:
            raise ValueError(f"第 {excel_row} 行坐标非法: start={start}, end={end}")
        if strand not in {"+", "-"}:
            raise ValueError(f"第 {excel_row} 行 strand 必须是 + 或 -: {strand!r}")
        if set(sequence) - IUPAC_BASES:
            bad = "".join(sorted(set(sequence) - IUPAC_BASES))
            raise ValueError(f"第 {excel_row} 行包含非法碱基字符: {bad}")
        if name in seen_names:
            raise ValueError(f"第 {excel_row} 行 primer 名称重复: {name}")
        seen_names.add(name)

        primer_on_reference = sequence if strand == "+" else reverse_complement(sequence)
        coordinate_adjustment = "none"
        if end - start != len(sequence):
            can_repair = (
                args.repair_one_based_inclusive
                and end - start + 1 == len(sequence)
                and chrom in reference
                and start >= 1
                and end <= len(reference[chrom])
                and count_mismatches(primer_on_reference, reference[chrom][start - 1 : end]) == 0
            )
            if can_repair:
                # 少数 patch primer 使用 1-based inclusive；经参考精确验证后转为 BED。
                start -= 1
                coordinate_adjustment = "start_minus_1_from_1_based_inclusive"
            else:
                raise ValueError(
                    f"第 {excel_row} 行坐标跨度 {end - start} 与序列长度 {len(sequence)} 不同；"
                    "输入坐标应为 BED 的 0-based、左闭右开格式"
                )

        mismatch_count: int | str = "NA"
        reference_status = "not_checked"
        if reference is not None:
            if chrom not in reference:
                raise ValueError(f"参考 FASTA 中找不到染色体名称: {chrom}")
            if end > len(reference[chrom]):
                raise ValueError(
                    f"第 {excel_row} 行终点 {end} 超过参考序列 {chrom} 长度 {len(reference[chrom])}"
                )
            target = reference[chrom][start:end]
            mismatch_count = count_mismatches(primer_on_reference, target)
            reference_status = "match" if mismatch_count == 0 else "mismatch"

        records.append(
            {
                "chrom": chrom,
                "start": start,
                "end": end,
                "name": name,
                "pool": pool,
                "strand": strand,
                "sequence": sequence,
                "input_start": input_start,
                "input_end": input_end,
                "coordinate_adjustment": coordinate_adjustment,
                "mismatch_count": mismatch_count,
                "reference_status": reference_status,
            }
        )

    if not records:
        raise ValueError("工作表中没有可转换的 primer 记录")

    # BED 必须按参考序列和坐标排序，iVar 才能稳定处理相邻及替代引物。
    records.sort(key=lambda item: (str(item["chrom"]), int(item["start"]), int(item["end"]), str(item["name"])))
    args.output_bed.parent.mkdir(parents=True, exist_ok=True)
    with args.output_bed.open("w", encoding="utf-8", newline="\n") as handle:
        writer = csv.writer(handle, delimiter="\t", lineterminator="\n")
        for record in records:
            writer.writerow(
                [
                    record["chrom"],
                    record["start"],
                    record["end"],
                    record["name"],
                    record["pool"],
                    record["strand"],
                ]
            )

    if args.audit_tsv:
        args.audit_tsv.parent.mkdir(parents=True, exist_ok=True)
        with args.audit_tsv.open("w", encoding="utf-8", newline="\n") as handle:
            fields = [
                "chrom",
                "start",
                "end",
                "name",
                "pool",
                "strand",
                "sequence",
                "input_start",
                "input_end",
                "coordinate_adjustment",
                "mismatch_count",
                "reference_status",
            ]
            writer = csv.DictWriter(handle, fieldnames=fields, delimiter="\t", lineterminator="\n")
            writer.writeheader()
            writer.writerows(records)

    mismatched = sum(record["reference_status"] == "mismatch" for record in records)
    repaired = sum(record["coordinate_adjustment"] != "none" for record in records)
    if args.strict_reference and mismatched:
        print(f"[ERROR] 有 {mismatched} 条 primer 与参考序列不完全匹配", file=sys.stderr)
        return 2

    print(f"[INFO] converted primers: {len(records)}")
    print(f"[INFO] BED: {args.output_bed}")
    if args.audit_tsv:
        print(f"[INFO] audit TSV: {args.audit_tsv}")
    if reference is not None:
        print(f"[INFO] exact reference matches: {len(records) - mismatched}")
        print(f"[INFO] primers with reference mismatches: {mismatched}")
    print(f"[INFO] repaired 1-based inclusive rows: {repaired}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (OSError, ValueError) as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        raise SystemExit(1)
